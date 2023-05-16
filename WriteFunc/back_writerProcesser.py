import openpyxl
import re
from utils.ExcelManager import ExcelManager

import sys
import time
import numpy as np
import copy

from PySide6 import QtWidgets,QtCore,QtGui
from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt,QPropertyAnimation,QEvent)
from PySide6.QtGui import *
    
from PySide6.QtWidgets import *
from .WriterEditorWindow import WriterEditor
from .ProcesserWindow  import Processer

from utils.PathManager import PathMgr
from .WriteDataManager import WriteDataManager
import utils.styleSheets as styles

class Processer():
    def __init__(self,writer,infoBoard,paths) -> None:
        self.writer = writer

        self.exlMgr = ExcelManager()
        self.wbs = []
        self.wbnames = []
        self.sheets = []
        self.sheetNames = []
        self.keys_dict = {}
        self.values_dict = {}
        self.widget = QWidget()
        self.infoBoard = infoBoard
        self.wbname_cache = None
        self.__private_addInfo("======="+writer.name+"运行=======")
        self.__private_GetExcels(paths,self.writer.workbookNames)
        self.wbPaths = self.exlMgr.GetWorkbookPaths()
        self.batchOprValue_dict = {}

        
    def Operate_batch(self,input_wb_path):
        input_wb = openpyxl.load_workbook(input_wb_path)
        input_sheet = input_wb.active
        length = 0
        for col in input_sheet.iter_cols():
            nameNow = None
            listNow = []
            for i,cell in enumerate(col):
                if i == 0:
                    nameNow = cell.value
                else:
                    listNow.append(cell.value)
            if len(listNow)>length:
                length = len(listNow)
            self.batchOprValue_dict[nameNow] = listNow

        

        for i in range(length):
            #refresh keys_dict and values_dict
            self.keys_dict.clear()
            self.values_dict.clear()
            for key,value in self.batchOprValue_dict.items():
                value = value[i]
                if key in self.writer.keyNames:
                    self.keys_dict[key] = value
                if key in self.writer.valueNames:
                    self.values_dict[key] = value
            if len(self.keys_dict) == 0:
                QMessageBox.warning(self.widget, "警告", "必须至少有一个检索变量有值")
                return
            self.__private_Operate()



    
    def Operate_Onece(self,k_list,v_list):
        
        if len(self.writer.workbookNames) == 0:
            QMessageBox.warning(self.widget, "警告", "需要填入该writer作用的工作簿名称")
            return
        
        self.__private_GetValueFromUI(k_list,v_list)

        self.__private_Operate()
        

    def __private_Operate(self):
        for i,wb in enumerate(self.wbs):
            written = False
            self.sheets,self.sheetNames = self.exlMgr.GetSheetsFromWorkbook(wb,self.writer.sheetNames)
            #now got sheet list
            for s in self.sheets:
                for process in self.writer.processes:
                    cord_list_tup = self.__private_GetTargetCellCordFromSheet(s,process.name)
                    if cord_list_tup is None or len(cord_list_tup) == 0:
                        continue
                    value = self.__private_GetProcessCalValue(process)
                    inputStr = process.inputStr
                    for key,value in self.values_dict:
                        if key in inputStr:
                            inputStr = inputStr.replace(key,str(value))
                    if "$" in inputStr:
                        inputStr = inputStr.replace("$",str(value))
                    for cord in cord_list_tup:
                        if not process.writeAll and written:
                            continue
                        #rewrite
                        if process.reWrite:
                            self.__private_WriteToCell(wb,s,cord,inputStr)
                        else:
                            existVal = str(s.cell(row=cord[0],column=cord[1]).value)
                            #自动添加等号
                            if inputStr[0] == "+":
                                if existVal is None or existVal[0] != "=":
                                    inputStr = inputStr.replace("+","=",1)
                                    existVal = ""
                                inputStr = existVal+inputStr
                            #自动补全前文末尾
                            if inputStr[len(inputStr)-1] in [";","."]:
                                if existVal[len(existVal)-1] != inputStr[len(inputStr)-1]:
                                    inputStr = existVal + inputStr[len(inputStr)-1] + inputStr
                            self.__private_WriteToCell(wb,s,cord,inputStr)
                        
                        written = True
            if written:
                wb.save(self.wbPaths[i])
            else:
                wbnameNow = self.wbnames[self.wbs.index(wb)]
                warning = wbnameNow+"中未找到录入对象"
                self.__private_addWarning(warning)

    def __private_WriteToCell(self,wb,sheet,cord_tup,value):
        wbnameNow = self.wbnames[self.wbs.index(wb)]
        if self.wbname_cache != wbnameNow:
            self.__private_addInfo("----在工作簿："+wbnameNow+" 中写入----")
            self.wbname_cache = wbnameNow
        
        sheetName = self.sheetNames[self.sheets.index(sheet)]
        sheet.cell(row = cord_tup[0],column = cord_tup[1]).value = value
        self.__private_addInfo("--->工作表："+sheetName)
        self.__private_addInfo("("+str(chr(cord_tup[1]+64))+str(cord_tup[0])+"):")
        value = self.insert_char_every_n_chars(value,"\n",20)
        self.__private_addInfo(value)


    def insert_char_every_n_chars(self,string, char, n):
        return char.join([string[i:i+n] for i in range(0, len(string), n)])

                    



    def __private_GetProcessCalValue(self,process):
        calStr = process.processStr
        if "\n" in calStr:
            calStr = calStr.replace("\n","")
        for key,value in self.values_dict.items():
            calStr = calStr.replace(key,str(value))
        try:
            value = float(eval(calStr))
        except:
            value = None
        if not type(value) == float:
            QMessageBox.warning(self.widget, "警告", "操作："+process.name+" 值计算有误，请检查名称是否与写入变量名称相匹配")
            
        return value




    def __private_GetTargetCellCordFromSheet(self,sheet,name):
        cellCord_list_tup = []
        if self.writer.isRowProcess:
            row_list = self.__private_GetTargetRowFromSheet(sheet)
            if row_list is None:
                return None
            for tup in row_list:
                range_list = tup[0]
                colIndex_list = tup[1]
                for row in sheet.iter_rows(min_row = 1,max_row = 1):
                    for ind in range_list:
                        if row[ind-1].value == name:
                            for rowind in colIndex_list:
                                cellCord_list_tup.append(copy.deepcopy((rowind,ind)))
            return cellCord_list_tup
        else:
            col_list = self.__private_GetTargetColFromSheet(sheet)
            for tup in col_list:
                range_list = tup[0]
                colIndex_list = tup[1]
                for col in sheet.iter_cols(min_col = 1,max_col = 1):
                    for ind in range_list:
                        if col[ind-1].value == name:
                            for colind in colIndex_list:
                                cellCord_list_tup.append((ind,colind))
            return cellCord_list_tup
        

    


    def __private_GetTargetRowFromSheet(self,sheet):
        #tup[0]->list:行作用域(只作用于列表中存在index的列)
        #tup[1]->list:该作用域内满足条件的行
        ROW_list_tup = [] 

        range = []
        rows = []

        masks = []
        found_name = []
        
        for row in sheet.iter_rows(min_row = 1,max_row = 1):
            for i,cell in enumerate(row):
                name = cell.value
                # if start a new row range?
                if name in found_name:
                    if len(masks) == 0:
                        return None
                    if self.writer.key_and_mode:
                        mask = np.logical_and.reduce(masks)
                    else:
                        mask = np.logical_or.reduce(masks)
                    
                    for ii,value in enumerate(mask):
                        if value:
                            rows.append(ii+1)
                    ROW_list_tup.append((copy.deepcopy(range),copy.deepcopy(rows)))

                    found_name.clear()
                    masks.clear()
                    range.clear()
                    rows.clear()
                    found_name.append(name)
                    range.append(i+1)

                elif i == len(row)-1:
                    found_name.append(name)
                    range.append(i+1)
                    if len(masks) == 0:
                        return None
                    if self.writer.key_and_mode:
                        mask = np.logical_and.reduce(masks)
                    else:
                        mask = np.logical_or.reduce(masks)
                    
                    for ii,value in enumerate(mask):
                        if value:
                            rows.append(ii+1)
                    ROW_list_tup.append((copy.deepcopy(range),copy.deepcopy(rows)))

                else:
                    found_name.append(name)
                    range.append(i+1)
                
                # if match the keyName
                if name in self.writer.keyNames and name in self.keys_dict.keys():
                    if self.keys_dict[name] is None:
                        continue
                    #get values of this col
                    for col in sheet.iter_cols(min_col = i+1,max_col = i+1):
                        value_mask = [c.value == self.keys_dict[name] for c in col]
                        masks.append(value_mask)

        return ROW_list_tup
    
    def __private_GetTargetColFromSheet(self,sheet):
        #tup[0]->list:列作用域(只作用于列表中存在index的列)
        #tup[1]->list:该作用域内满足条件的行
        COL_list_tup = [] 

        range = []
        rows = []

        masks = []
        found_name = []
        
        for col in sheet.iter_cols(min_col = 1,max_col = 1):
            for i,cell in enumerate(col):
                name = cell.value
                # if start a new col range?
                if name in found_name :

                    if len(masks) == 0:
                        return None
                    if self.writer.key_and_mode:
                        mask = np.logical_and.reduce(masks)
                    else:
                        mask = np.logical_or.reduce(masks)
                    for ii,value in enumerate(mask):
                        if value:
                            rows.append(ii+1)
                    COL_list_tup.append((copy.deepcopy(range),copy.deepcopy(rows)))

                    found_name.clear()
                    masks.clear()
                    range.clear()
                    rows.clear()
                    found_name.append(name)
                    range.append(i+1)

                elif i == len(col)-1:
                    found_name.append(name)
                    range.append(i+1)

                    if len(masks) == 0:
                        return None

                    if self.writer.key_and_mode:
                        mask = np.logical_and.reduce(masks)
                    else:
                        mask = np.logical_or.reduce(masks)
                    for ii,value in enumerate(mask):
                        if value:
                            rows.append(ii+1)
                    COL_list_tup.append((copy.deepcopy(range),copy.deepcopy(rows)))


                else:
                    found_name.append(name)
                    range.append(i+1)
                
                # if match the keyName
                if name in self.writer.keyNames and self.keys_dict[name] is not None:
                    #get values of this col
                    for row in sheet.iter_rows(min_row = i+1,max_row = i+1):
                        value_mask = [c.value == self.keys_dict[name] for c in row]
                        masks.append(value_mask)

        return COL_list_tup



    def __private_GetExcels(self,paths,keywords):
        self.wbs,self.wbnames = self.exlMgr.GetExcelWorkBook(paths,keywords)
        self.__private_addInfo("找到以下工作簿：")
        for n in self.wbnames:
            self.__private_addInfo(n)
        self.__private_addInfo("--------------------")

        


    def __private_GetValueFromUI(self,key_list,value_list):
        for key in self.writer.keyNames:
            for line in key_list:
                if line.objectName() == key:
                    number = line.text()
                    if number == "":
                        self.keys_dict[key] = None
                        break
                    number = number.replace(" ","")
                    if not self.__private_Is_number(number):
                        QMessageBox.warning(self.widget, "警告", "填入值不能有除数字字符外的其他符号，包括空格、*等")
                        break
                    else:
                        number = float(number)
                    self.keys_dict[key] = number
                    break
        for value in self.writer.valueNames:
            for line in value_list:
                if line.objectName() == value:
                    number = line.text()
                    if number == "":
                        self.values_dict[value] = None
                        break
                    number = number.replace(" ","")
                    if not self.__private_Is_number(number):
                        QMessageBox.warning(self.widget,"警告", "填入值不能有除数字字符外的其他符号，包括空格、*等")
                        break
                    else:
                        number = float(number)
                    self.values_dict[value] = number
                    break



    def __private_Is_number(self,str):
        pattern = re.compile(r'^[-+]?[0-9]+\.[0-9]+$|^[-+]?[0-9]+$')
        return pattern.match(str) is not None
    
    def __private_addInfo(self,info):
        self.infoBoard.addItem(QListWidgetItem(info))

    def __private_addWarning(self,warning):
        item = QListWidgetItem(warning)
        item.setForeground(QColor(255,0,0))
        self.infoBoard.addItem(item)
    


    