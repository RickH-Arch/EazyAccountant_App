import re
import os
from openpyxl import load_workbook,Workbook


class ExcelManager:
    def __init__(self) -> None:
        self.workbookCache = []
        self.workbookNameCache = []
        self.workbookPathCache = []

    def GetWorkbookPaths(self):
        return self.workbookPathCache

    def GetExcelWorkBook(self,path,words):
        '''
        return workbooks as list, and name of each workbook as list
        '''
        self.__private_ResetCache()
        if isinstance(path,list):
            workbooks,names = self.__private_LoadExcelWorkBookFromPaths(path,words)
        else:
            workbooks,names = self.__private_LoadExcelWorkBookFromPath(path,words)
        return workbooks,names
    
    def GetSheetsFromWorkbook(self,workbook,keywords):
        '''
        get sheets from one workbook
        return sheets as list, and name of each sheet as list
        '''
        sheets = []
        sheetNames = []
        sheet_names = workbook.sheetnames
        for sheetName in sheet_names:
            if len(keywords) != 0:
                get = False
                for keyword in keywords:
                    if self.AnalyzeFileName(sheetName,keyword):
                        get = True
                if get:
                    sheets.append(workbook[sheetName])
                    sheetNames.append(sheetName)
            else:
                sheets.append(workbook[sheetName])
                sheetNames.append(sheetName)
        return sheets,sheetNames
    
    def GetValueFromSheet(self,sheet,cord):
        c = self.__private_StringToCord(cord)
        cell = sheet.cell(c[0],c[1])
        value = cell.value
        if value == "":
            value = "N/A"
        return value
    
    def SortWorkbookByName(self,workbooks,workbookNames):
        dates = []
        for n in workbookNames:
            ds = re.findall('\d+',n)
            d = int(''.join(map(str,ds)))
            dates.append(d)
        zipped = zip(dates,workbooks)
        zippedName = zip(dates,workbookNames)
        sortrd_pairs = sorted(zipped)
        sorted_pairs_name = sorted(zippedName)
        ws = [pair[1] for pair in sortrd_pairs]
        wns = [pair[1] for pair in sorted_pairs_name]
        return ws,wns
    
    def CreateExcel(self,colHeaders,rowHeaders):
        wb = Workbook()
        ws = wb.create_sheet("sheet1",0)
        colCur = 2
        for colHeader in colHeaders:
            ws.cell(row = 1,column = colCur,value = colHeader)
            colCur+=1
        rowCur = 2
        for rowHeader in rowHeaders:
            ws.cell(row = rowCur,column = 1,value = rowHeader)
            rowCur+=1
        return wb

    
    
    def __private_LoadExcelWorkBookFromPath(self,folder_path,file_keywords):
        
        for file_name in os.listdir(folder_path):
            file_path = os.path.join(folder_path,file_name)
            if os.path.isdir(file_path):
                self.__private_LoadExcelWorkBookFromPath(file_path,file_keywords)
            elif file_name.endswith(".xlsx") or file_name.endswith(".xls"):
                if len(file_keywords)!=0:
                    getKeyword = False
                    for w in file_keywords:
                        if self.AnalyzeFileName(file_name,w):
                            getKeyword = True
                    if file_name not in self.workbookNameCache and getKeyword:
                        workbook = load_workbook(file_path)
                        self.workbookCache.append(workbook)
                        self.workbookNameCache.append(file_name)
                        self.workbookPathCache.append(file_path)
                elif file_name not in self.workbookNameCache:
                    workbook = load_workbook(file_path)
                    self.workbookCache.append(workbook)
                    self.workbookNameCache.append(file_name)
                    self.workbookPathCache.append(file_path)

        return self.workbookCache,self.workbookNameCache

    
    def __private_LoadExcelWorkBookFromPaths(self,paths,file_keywords):
        
        for path in paths:
            self.__private_LoadExcelWorkBookFromPath(path,file_keywords)
        return self.workbookCache,self.workbookNameCache
    
    
    def __private_ResetCache(self):
        self.workbookCache.clear()
        self.workbookNameCache.clear()

    



    
    def AnalyzeFileName(self,fileName,name):
        '''
        return True if fileName contains name
        '''
        ind = 0
        get = False
        start = False
        if '*' in name:
            for i,w in enumerate(fileName) :
                if ind>=len(name):
                    continue
                if name[ind] == '*':
                    ind += 1
                    continue
                if w == name[ind]:
                    start = True
                    get = True
                    ind += 1
                elif(start):
                    get = False
                    start = False
                    ind = 0
        elif name in fileName:
            get = True
        return get

    def __private_StringToCord(self,str_):

        if type(str_) == str:
            cord = [int(str_[1:]) , ord(str_[0]) - 64]
            return cord
        else:
          return [-1,-1]
  

    
